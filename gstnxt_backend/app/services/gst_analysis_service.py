import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional
from datetime import datetime
import os
from pathlib import Path
import uuid
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from app.models import GSTProject, FileUpload, AnalysisResult
from decouple import config

class GSTAnalysisService:
    """Service for analyzing GST data and generating reports"""
    
    # Worksheet configurations as per requirements
    GSTR1_WORKSHEETS = {
        'B2B': {
            'reference': 'Taxable supplies made to registered taxpayers (invoices only) - B2B',
            'table_ref': '4A, 4B, 4C, 6B, 6C',
            'output_name': 'Outward - b2b+sez+de',
            'month_column': 'R',
            'date_column': 'D',
            'validation_required': True
        },
        'SEZ': {
            'reference': 'Taxable supplies made to registered taxpayers (invoices only) - SEZ',
            'table_ref': '4A, 4B, 4C, 6B, 6C',
            'output_name': 'Outward - b2b+sez+de',
            'month_column': 'R',
            'date_column': 'D',
            'validation_required': True
        },
        'DE': {
            'reference': 'Taxable supplies made to registered taxpayers (invoices only) - DE',
            'table_ref': '4A, 4B, 4C, 6B, 6C',
            'output_name': 'Outward - b2b+sez+de',
            'month_column': 'R',
            'date_column': 'D',
            'validation_required': True
        },
        'B2CL': {
            'reference': 'Taxable outward inter-state supplies made to unregistered persons (where invoice value is more than Rs.2.5 lakh) - B2CL (Large)',
            'table_ref': '5',
            'output_name': 'Out - b2cl',
            'month_column': 'I',
            'date_column': 'A',
            'validation_required': True
        },
        'EXP': {
            'reference': 'Exports',
            'table_ref': '6A',
            'output_name': 'Out - exp',
            'month_column': 'O',
            'date_column': 'C',
            'validation_required': True
        },
        'B2CS': {
            'reference': 'Taxable supplies (Net of debit and credit notes) to unregistered persons (other than the supplies covered in Table 5) - B2CS (Others)',
            'table_ref': '7',
            'output_name': 'Out-b2cs',
            'month_column': None,
            'date_column': None,
            'validation_required': False
        },
        'EXEMP': {
            'reference': 'Nil rated, exempted and non GST outward supplies',
            'table_ref': '8',
            'output_name': 'Out - exemp',
            'month_column': None,
            'date_column': None,
            'validation_required': False
        },
        'B2BA': {
            'reference': 'Amendment to taxable outward supplies made to registered person in returns of earlier tax periods in table 4A, 4B, 6B, 6C - B2B',
            'table_ref': '9A',
            'output_name': 'Out - b2ba',
            'month_column': None,
            'date_column': None,
            'validation_required': False
        },
        'B2CLA': {
            'reference': 'Amendment to Inter-State supplies made to unregistered person (where invoice value is more than Rs.2.5 lakh) in returns of earlier tax periods in table 5 - B2CL (Large)',
            'table_ref': '9B',
            'output_name': 'Out - b2cla',
            'month_column': None,
            'date_column': None,
            'validation_required': False
        },
        'EXPA': {
            'reference': 'Amendment to Export Supplies',
            'table_ref': '9C',
            'output_name': 'Out - expa',
            'month_column': None,
            'date_column': None,
            'validation_required': False
        },
        'CDNR': {
            'reference': 'Credit/ Debit notes issued to the registered taxpayers - CDNR',
            'table_ref': '10',
            'output_name': 'Outward - cdnr',
            'month_column': 'S',
            'date_column': 'D',
            'validation_required': True
        },
        'CDNUR': {
            'reference': 'Credit/ Debit notes issued to the unregistered persons - CDNUR',
            'table_ref': '11',
            'output_name': 'Out - cdnur',
            'month_column': 'N',
            'date_column': 'C',
            'validation_required': True
        },
        'CDNRA': {
            'reference': 'Amendment to Credit/ Debit notes issued to the registered taxpayers - CDNRA',
            'table_ref': '12A',
            'output_name': 'Out-cdnra',
            'month_column': 'R',
            'date_column': None,
            'validation_required': False
        },
        'CDNURA': {
            'reference': 'Amendment to Credit/ Debit notes issued to the unregistered persons - CDNURA',
            'table_ref': '12B',
            'output_name': 'Out-cdnura',
            'month_column': 'M',
            'date_column': None,
            'validation_required': False
        },
        'B2CSA': {
            'reference': 'Amendment to taxable outward supplies made to unregistered person in returns for earlier tax periods in table 7 - B2C (Others)',
            'table_ref': '12C',
            'output_name': 'Out-b2csa',
            'month_column': 'J',
            'date_column': None,
            'validation_required': False
        },
        'AT': {
            'reference': '11A(1), 11A(2) - Advances received for which invoice has not been issued (tax amount to be added to the output tax liability) (Net of refund vouchers)',
            'table_ref': '11A(1), 11A(2)',
            'output_name': 'Out-at',
            'month_column': 'H',
            'date_column': None,
            'validation_required': False
        },
        'ATADJ': {
            'reference': '11B(1), 11B(2) - Advance amount received in earlier tax period and adjusted against the supplies being shown in this tax period in Table Nos. 4, 5, 6 and 7',
            'table_ref': '11B(1), 11B(2)',
            'output_name': 'Out-atadj',
            'month_column': 'H',
            'date_column': None,
            'validation_required': False
        },
        'ATA': {
            'reference': '11A - Amendment to advances received in returns for earlier tax periods in table 11A(1), 11A(2)',
            'table_ref': '11A',
            'output_name': 'Out-ata',
            'month_column': 'J',
            'date_column': None,
            'validation_required': False
        },
        'ATADJA': {
            'reference': '11B - Amendment to advances adjusted in returns for earlier tax periods in table 11B(1), 11B(2)',
            'table_ref': '11B',
            'output_name': 'Out-atadja',
            'month_column': 'J',
            'date_column': None,
            'validation_required': False
        },
        'HSN': {
            'reference': '12 - HSN-wise summary of outward supplies',
            'table_ref': '12',
            'output_name': 'Out-hsn',
            'month_column': 'K',
            'date_column': None,
            'validation_required': False
        },
        'DOCS': {
            'reference': '13 - Documents issued',
            'table_ref': '13',
            'output_name': 'Out-docs',
            'month_column': 'F',
            'date_column': None,
            'validation_required': False
        }
    }
    
    GSTR2A_WORKSHEETS = {
        'B2B': {
            'reference': 'Taxable inward supplies received from registered person',
            'table_ref': '3'
        },
        'B2BA': {
            'reference': 'Amendments to previously uploaded invoices by supplier',
            'table_ref': '4'
        },
        'CDNR': {
            'reference': 'Debit/Credit notes(Original)',
            'table_ref': '5'
        },
        'CDNRA': {
            'reference': 'Amendments to previously uploaded Credit/Debit notes by supplier',
            'table_ref': '6'
        },
        'ECO': {
            'reference': 'Documents reported by ECO on which ECO is liable to pay tax u/s 9(5)',
            'table_ref': '7'
        },
        'ECOA': {
            'reference': 'Amendment to documents reported by ECO on which ECO is liable to pay tax u/s 9(5)',
            'table_ref': '8'
        },
        'ISD': {
            'reference': 'ISD Credit',
            'table_ref': '9'
        },
        'ISDA': {
            'reference': 'Amendments to iSD Credits received',
            'table_ref': '10'
        },
        'TDS': {
            'reference': 'TDS Credit received',
            'table_ref': '11'
        },
        'TDSA': {
            'reference': 'Amendments to TDS Credit received',
            'table_ref': '12'
        },
        'TCS': {
            'reference': 'Details of supplies made through e-commerce operator (TCS)',
            'table_ref': '13'
        },
        'TCSA': {
            'reference': 'Amendments to details of supplies in respect of any earlier statement (TCSA)',
            'table_ref': '14'
        },
        'IMPG': {
            'reference': 'Import of goods from overseas on bill of entry',
            'table_ref': '15'
        },
        'IMPGSEZ': {
            'reference': 'Import of goods from SEZ units/developers on bill of entry',
            'table_ref': '16'
        }
    }
    
    @staticmethod
    def start_analysis(db: Session, project_id: str) -> Dict[str, Any]:
        """Start GST data analysis for a project"""
        try:
            # Get project details
            project = db.query(GSTProject).filter(GSTProject.id == project_id).first()
            if not project:
                return {"success": False, "error": "Project not found"}
            
            # Get uploaded files
            uploads = db.query(FileUpload).filter(
                FileUpload.project_id == project_id,
                FileUpload.upload_status == 'uploaded'
            ).all()
            
            if not uploads:
                return {"success": False, "error": "No uploaded files found"}
            
            # Create analysis result record
            analysis = AnalysisResult(
                project_id=project_id,
                analysis_type='comparison',
                status='processing'
            )
            db.add(analysis)
            db.commit()
            db.refresh(analysis)
            
            # Analyze data
            analysis_result = GSTAnalysisService._process_gst_data(uploads, project, str(analysis.id))
            
            if analysis_result["success"]:
                # Update analysis record
                analysis.status = 'completed'
                analysis.output_filename = analysis_result["output_filename"]
                analysis.output_file_path = analysis_result["output_path"]
                analysis.analysis_summary = analysis_result["summary"]
                analysis.completed_at = datetime.utcnow()
            else:
                analysis.status = 'failed'
                analysis.error_message = analysis_result["error"]
            
            db.commit()
            
            return {
                "success": analysis_result["success"],
                "analysis_id": str(analysis.id),
                "output_filename": analysis_result.get("output_filename"),
                "summary": analysis_result.get("summary"),
                "error": analysis_result.get("error")
            }
            
        except Exception as e:
            return {"success": False, "error": f"Analysis failed: {str(e)}"}
    
    @staticmethod
    def _process_gst_data(uploads: List[FileUpload], project: GSTProject, analysis_id: str) -> Dict[str, Any]:
        """Process GST data and create output Excel file"""
        try:
            # Group uploads by file type
            gstr1_files = [u for u in uploads if u.file_type == 'GSTR1']
            gstr2a_files = [u for u in uploads if u.file_type == 'GSTR2A']
            
            # Create output directory
            output_dir = Path(config('OUTPUT_DIR', default='outputs'))
            output_dir.mkdir(exist_ok=True)
            
            # Generate output filename
            output_filename = f"GST_Analysis_{project.gstin}_{project.financial_year}_{analysis_id[:8]}.xlsx"
            output_path = output_dir / output_filename
            
            # Create Excel workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Track created worksheets
            created_worksheets = []
            
            # Process GSTR1 data first to get list of created worksheets
            gstr1_summary = GSTAnalysisService._process_gstr1_data(wb, gstr1_files)
            created_worksheets.extend([ws.title for ws in wb.worksheets if ws.title != 'Index'])
            
            # Process GSTR2A data
            gstr2a_summary = GSTAnalysisService._process_gstr2a_data(wb, gstr2a_files)
            # Add any new worksheets created by GSTR2A processing
            all_worksheet_names = [ws.title for ws in wb.worksheets if ws.title != 'Index']
            for name in all_worksheet_names:
                if name not in created_worksheets:
                    created_worksheets.append(name)
            
            # Create index sheet with accurate information about created worksheets and analysis details
            gstr1_analysis_details = gstr1_summary.get("analysis_details", {})
            index_data = GSTAnalysisService._create_index_sheet_data(gstr1_files, gstr2a_files, created_worksheets, gstr1_analysis_details)
            GSTAnalysisService._add_index_sheet(wb, index_data)
            
            # Save workbook
            wb.save(output_path)
            
            print(f"Analysis complete. Created {len(created_worksheets)} data worksheets: {created_worksheets}")
            
            return {
                "success": True,
                "output_filename": output_filename,
                "output_path": str(output_path),
                "summary": {
                    "gstr1_summary": gstr1_summary,
                    "gstr2a_summary": gstr2a_summary,
                    "total_worksheets": len(wb.worksheets),
                    "created_data_worksheets": created_worksheets
                }
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"Data processing failed: {str(e)}"
            }
    
    @staticmethod
    def _create_index_sheet_data(gstr1_files: List[FileUpload], gstr2a_files: List[FileUpload], created_worksheets: List[str], analysis_details: Dict = None) -> List[Dict]:
        """Create index sheet data based on available files and actually created worksheets"""
        index_data = []
        
        # GSTR1 worksheets
        for i, (worksheet_name, config) in enumerate(GSTAnalysisService.GSTR1_WORKSHEETS.items(), 1):
            # Check if worksheet was actually created or combined
            output_name = config['output_name']
            analysis_output = "Taxpayer has no records under this head. Hence, No analysis is required"
            remarks = ""
            
            if created_worksheets and (worksheet_name in created_worksheets or output_name in created_worksheets):
                analysis_output = "Analysed and Generated"
                
                # Add remarks for wrong month records if available
                if analysis_details and worksheet_name in analysis_details:
                    wrong_count = analysis_details[worksheet_name].get("wrong_month_records", 0)
                    if wrong_count > 0:
                        remarks = f"{wrong_count} wrong month invoices/rows"
            
            index_data.append({
                "S. No": i,
                "Worksheet Name": worksheet_name,
                "GSTR-1 Table Reference": config['reference'],
                "Output Sheet Name": output_name,
                "Analysis Output": analysis_output,
                "Remarks": remarks
            })
        
        # GSTR2A worksheets
        for i, (worksheet_name, config) in enumerate(GSTAnalysisService.GSTR2A_WORKSHEETS.items(), 1):
            # Check if worksheet was actually created
            analysis_output = "Analysed and Generated" if created_worksheets and worksheet_name in created_worksheets else "Taxpayer has no records under this head. Hence, No analysis is required"
            
            index_data.append({
                "S. No (2A)": i,
                "Worksheet Name (2A)": worksheet_name,
                "GSTR-2A Table Reference": config['reference'],
                "Analysis Output (2A)": analysis_output
            })
        
        return index_data
    
    @staticmethod
    def _add_index_sheet(wb: Workbook, index_data: List[Dict]):
        """Add formatted index sheet to workbook with professional table styling"""
        ws = wb.create_sheet("Index", 0)
        
        # Define professional color scheme
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue header
        alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray alternating rows
        header_font = Font(color="FFFFFF", bold=True, size=11)  # White bold text for headers
        normal_font = Font(color="000000", size=10)  # Black text for data
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add GSTR-1 headers
        gstr1_headers = ["S. No", "Worksheet Name", "GSTR-1 Table Reference", "Analysis Output"]
        for col, header in enumerate(gstr1_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Add empty column
        ws.cell(row=1, column=5, value="")
        
        # Add GSTR-2A headers
        gstr2a_headers = ["S. No", "Worksheet Name", "GSTR-2A Table Reference", "Analysis Output"]
        for col, header in enumerate(gstr2a_headers, 6):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Add data rows with alternating colors
        max_rows = max(len(GSTAnalysisService.GSTR1_WORKSHEETS), len(GSTAnalysisService.GSTR2A_WORKSHEETS))
        
        for i in range(max_rows):
            row_num = i + 2
            use_alt_color = (i % 2 == 1)  # Alternate every other row
            
            # GSTR1 data
            if i < len(GSTAnalysisService.GSTR1_WORKSHEETS):
                gstr1_items = list(GSTAnalysisService.GSTR1_WORKSHEETS.items())
                worksheet_name, config = gstr1_items[i]
                
                # S. No
                cell = ws.cell(row=row_num, column=1, value=i + 1)
                cell.font = normal_font
                cell.alignment = center_alignment
                cell.border = border
                if use_alt_color:
                    cell.fill = alt_row_fill
                
                # Worksheet Name
                cell = ws.cell(row=row_num, column=2, value=worksheet_name)
                cell.font = normal_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
                if use_alt_color:
                    cell.fill = alt_row_fill
                
                # GSTR-1 Table Reference
                cell = ws.cell(row=row_num, column=3, value=config['reference'])
                cell.font = normal_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
                if use_alt_color:
                    cell.fill = alt_row_fill
                
                # Analysis Output
                cell = ws.cell(row=row_num, column=4, value="Analysed and Generated")
                cell.font = normal_font
                cell.alignment = center_alignment
                cell.border = border
                if use_alt_color:
                    cell.fill = alt_row_fill
            
            # Empty column
            ws.cell(row=row_num, column=5, value="")
            
            # GSTR2A data
            if i < len(GSTAnalysisService.GSTR2A_WORKSHEETS):
                gstr2a_items = list(GSTAnalysisService.GSTR2A_WORKSHEETS.items())
                worksheet_name, config = gstr2a_items[i]
                
                # S. No
                cell = ws.cell(row=row_num, column=6, value=i + 1)
                cell.font = normal_font
                cell.alignment = center_alignment
                cell.border = border
                if use_alt_color:
                    cell.fill = alt_row_fill
                
                # Worksheet Name
                cell = ws.cell(row=row_num, column=7, value=worksheet_name)
                cell.font = normal_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
                if use_alt_color:
                    cell.fill = alt_row_fill
                
                # GSTR-2A Table Reference
                cell = ws.cell(row=row_num, column=8, value=config['reference'])
                cell.font = normal_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
                if use_alt_color:
                    cell.fill = alt_row_fill
                
                # Analysis Output
                cell = ws.cell(row=row_num, column=9, value="Analysed and Generated")
                cell.font = normal_font
                cell.alignment = center_alignment
                cell.border = border
                if use_alt_color:
                    cell.fill = alt_row_fill
        
        # Adjust column widths for better readability
        ws.column_dimensions['A'].width = 8   # S. No
        ws.column_dimensions['B'].width = 15  # Worksheet Name
        ws.column_dimensions['C'].width = 50  # Table Reference
        ws.column_dimensions['D'].width = 20  # Analysis Output
        ws.column_dimensions['E'].width = 3   # Separator
        ws.column_dimensions['F'].width = 8   # S. No
        ws.column_dimensions['G'].width = 15  # Worksheet Name
        ws.column_dimensions['H'].width = 50  # Table Reference
        ws.column_dimensions['I'].width = 20  # Analysis Output
    
    @staticmethod
    def _process_gstr1_data(wb: Workbook, gstr1_files: List[FileUpload]) -> Dict[str, Any]:
        """Process GSTR1 data with comprehensive validation and chronological sorting"""
        summary = {"files_processed": len(gstr1_files), "worksheets_created": 0, "analysis_details": {}}
        
        if not gstr1_files:
            return summary
        
        # Sort files chronologically (oldest first) by financial year order
        def get_fy_month_order(upload):
            """Convert calendar month to financial year month order"""
            if upload.month >= 4:  # Apr to Dec
                return upload.month - 3  # Apr=1, May=2, ..., Dec=9
            else:  # Jan to Mar
                return upload.month + 9  # Jan=10, Feb=11, Mar=12
        
        sorted_uploads = sorted(gstr1_files, key=lambda x: (x.year if x.month >= 4 else x.year + 1, get_fy_month_order(x)))
        
        # Group sheets that should be combined (B2B, SEZ, DE -> Outward - b2b+sez+de)
        combined_sheets = {}
        
        # Process each worksheet type
        for worksheet_name, config in GSTAnalysisService.GSTR1_WORKSHEETS.items():
            try:
                print(f"Processing GSTR1 worksheet: {worksheet_name}")
                
                all_monthly_data = []
                wrong_month_count = 0
                files_with_data = 0
                
                # Read data from all monthly files for this worksheet
                for upload in sorted_uploads:
                    monthly_data = GSTAnalysisService._read_worksheet_data(upload.file_path, worksheet_name)
                    
                    if monthly_data is not None and not monthly_data.empty:
                        # Check if there's actual data (not just headers)
                        data_rows = monthly_data.dropna(how='all')
                        if len(data_rows) > 0:
                            files_with_data += 1
                            
                            # Add month validation if required
                            if config.get('validation_required') and config.get('month_column') and config.get('date_column'):
                                monthly_data, wrong_count = GSTAnalysisService._add_month_validation(
                                    monthly_data, config['month_column'], config['date_column']
                                )
                                wrong_month_count += wrong_count
                            
                            all_monthly_data.append(monthly_data)
                
                # Combine all monthly data for this worksheet
                if all_monthly_data:
                    # Combine data chronologically (oldest first)
                    combined_data = pd.concat(all_monthly_data, ignore_index=True)
                    
                    # Handle combined sheets (B2B, SEZ, DE should go to same output)
                    output_name = config['output_name']
                    if output_name == 'Outward - b2b+sez+de':
                        if output_name not in combined_sheets:
                            combined_sheets[output_name] = []
                        combined_sheets[output_name].append(combined_data)
                    else:
                        # Create individual worksheet
                        GSTAnalysisService._create_formatted_worksheet(wb, output_name, combined_data, config)
                        summary["worksheets_created"] += 1
                    
                    # Track analysis details
                    summary["analysis_details"][worksheet_name] = {
                        "status": "Analysed and Generated",
                        "files_with_data": files_with_data,
                        "total_records": len(combined_data),
                        "wrong_month_records": wrong_month_count
                    }
                    
                    print(f"Processed worksheet {worksheet_name} with {len(combined_data)} records")
                    if wrong_month_count > 0:
                        print(f"Found {wrong_month_count} wrong month records in {worksheet_name}")
                else:
                    # No data found
                    summary["analysis_details"][worksheet_name] = {
                        "status": "Taxpayer has no records under this head. Hence, No analysis is required",
                        "files_with_data": 0,
                        "total_records": 0,
                        "wrong_month_records": 0
                    }
                    print(f"Skipped worksheet: {worksheet_name} (no data found)")
                    
            except Exception as e:
                print(f"Error processing worksheet {worksheet_name}: {e}")
                summary["analysis_details"][worksheet_name] = {
                    "status": "Error in processing",
                    "files_with_data": 0,
                    "total_records": 0,
                    "wrong_month_records": 0
                }
                continue
        
        # Create combined sheets (B2B+SEZ+DE)
        for output_name, data_list in combined_sheets.items():
            if data_list:
                combined_data = pd.concat(data_list, ignore_index=True)
                config = next(c for c in GSTAnalysisService.GSTR1_WORKSHEETS.values() 
                             if c['output_name'] == output_name)
                GSTAnalysisService._create_formatted_worksheet(wb, output_name, combined_data, config)
                summary["worksheets_created"] += 1
                print(f"Created combined worksheet {output_name} with {len(combined_data)} total records")
        
        return summary
    
    @staticmethod
    def _process_gstr2a_data(wb: Workbook, gstr2a_files: List[FileUpload]) -> Dict[str, Any]:
        """Process GSTR2A data and create worksheets only when data exists"""
        summary = {"files_processed": len(gstr2a_files), "worksheets_created": 0}
        
        if not gstr2a_files:
            return summary
        
        # Process each worksheet type
        for worksheet_name, config in GSTAnalysisService.GSTR2A_WORKSHEETS.items():
            try:
                # Collect data from all months chronologically (Apr to Mar)
                combined_data = []
                has_any_data = False
                
                # Sort files chronologically by financial year month order (Apr=1, May=2, ..., Mar=12)
                def get_fy_month_order(upload):
                    """Convert calendar month to financial year month order"""
                    if upload.month >= 4:  # Apr to Dec
                        return upload.month - 3  # Apr=1, May=2, ..., Dec=9
                    else:  # Jan to Mar
                        return upload.month + 9  # Jan=10, Feb=11, Mar=12
                
                sorted_uploads = sorted(gstr2a_files, key=lambda x: (x.year if x.month >= 4 else x.year + 1, get_fy_month_order(x)))
                
                for upload in sorted_uploads:
                    monthly_data = GSTAnalysisService._read_worksheet_data(upload.file_path, worksheet_name)
                    if monthly_data is not None and not monthly_data.empty and len(monthly_data) > 0:
                        # Check if there's actual data (not just headers)
                        data_rows = monthly_data.dropna(how='all')
                        if len(data_rows) > 0:
                            has_any_data = True
                            # Add month column at the end
                            month_name = GSTAnalysisService._get_month_name(upload.month, upload.year)
                            monthly_data['Month'] = month_name
                            combined_data.append(monthly_data)
                
                # Only create worksheet if there's actual data
                if has_any_data and combined_data:
                    # Combine all monthly data
                    final_data = pd.concat(combined_data, ignore_index=True)
                    
                    # Create worksheet with compiled data
                    GSTAnalysisService._create_formatted_worksheet(wb, worksheet_name, final_data, config)
                    summary["worksheets_created"] += 1
                    print(f"Created worksheet: {worksheet_name} with {len(final_data)} rows")
                else:
                    print(f"Skipped worksheet: {worksheet_name} (no data found)")
                
            except Exception as e:
                print(f"Error processing worksheet {worksheet_name}: {e}")
                continue
        
        return summary
    
    @staticmethod
    def _read_worksheet_data(file_path: str, worksheet_name: str) -> Optional[pd.DataFrame]:
        """Read data from specific worksheet in Excel file or ZIP containing Excel files, skipping header rows"""
        try:
            # Handle ZIP files
            if file_path.lower().endswith('.zip'):
                import zipfile
                import tempfile
                import os
                
                all_data = []
                
                with zipfile.ZipFile(file_path, 'r') as zip_ref:
                    # Find Excel files in ZIP
                    excel_files = [f for f in zip_ref.namelist() 
                                 if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('__MACOSX')]
                    
                    for excel_file in excel_files:
                        temp_file_path = None
                        try:
                            # Create temporary file with unique name
                            fd, temp_file_path = tempfile.mkstemp(suffix='.xlsx')
                            
                            # Write Excel data to temp file
                            with os.fdopen(fd, 'wb') as temp_file:
                                temp_file.write(zip_ref.read(excel_file))
                            
                            # Try to read the worksheet from this Excel file, skipping first 4 rows
                            data = GSTAnalysisService._read_excel_worksheet(temp_file_path, worksheet_name, skip_header_rows=True)
                            
                            if data is not None and not data.empty:
                                # Extract period from filename
                                period = GSTAnalysisService._extract_period_from_filename(excel_file)
                                # Add Month column as the last column only if it doesn't exist
                                if 'Month' not in data.columns:
                                    data['Month'] = period
                                all_data.append(data)
                                
                        except Exception as e:
                            print(f"Error reading Excel file {excel_file} from ZIP: {e}")
                            continue
                        finally:
                            # Clean up temp file
                            if temp_file_path and os.path.exists(temp_file_path):
                                try:
                                    os.unlink(temp_file_path)
                                except:
                                    pass
                
                # Combine all data from the ZIP
                if all_data:
                    return pd.concat(all_data, ignore_index=True)
                else:
                    return None
            
            # Handle direct Excel files
            else:
                data = GSTAnalysisService._read_excel_worksheet(file_path, worksheet_name, skip_header_rows=True)
                if data is not None and not data.empty:
                    # Extract period from filename
                    filename = os.path.basename(file_path)
                    period = GSTAnalysisService._extract_period_from_filename(filename)
                    # Add Month column as the last column only if it doesn't exist
                    if 'Month' not in data.columns:
                        data['Month'] = period
                return data
                
        except Exception as e:
            print(f"Error reading worksheet {worksheet_name} from {file_path}: {e}")
            return None
    
    @staticmethod
    def _read_excel_worksheet(file_path: str, worksheet_name: str, skip_header_rows: bool = False) -> Optional[pd.DataFrame]:
        """Read data from specific worksheet in Excel file"""
        excel_file = None
        try:
            excel_file = pd.ExcelFile(file_path)
            
            # Try to find the worksheet (case-insensitive and partial match)
            sheet_name = None
            for sheet in excel_file.sheet_names:
                # Exact match (case-insensitive)
                if sheet.upper() == worksheet_name.upper():
                    sheet_name = sheet
                    break
                # Partial match for cases like "B2B" matching "B2B (4A)"
                elif worksheet_name.upper() in sheet.upper() or sheet.upper() in worksheet_name.upper():
                    sheet_name = sheet
                    break
            
            if sheet_name:
                # Read the Excel sheet
                if skip_header_rows:
                    # For GSTR1 data, we need to:
                    # 1. Skip first 3 rows (title and subtitle)
                    # 2. Use row 4 as header (0-indexed row 3)
                    # 3. Start data from row 5 (0-indexed row 4)
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=3, skiprows=0)
                else:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                # Clean the data: remove completely empty rows and columns
                df = df.dropna(how='all')  # Remove rows where all values are NaN
                df = df.dropna(axis=1, how='all')  # Remove columns where all values are NaN
                
                # Clean column names - remove any leading/trailing whitespace and unnamed columns
                df.columns = [col.strip() if isinstance(col, str) and not col.startswith('Unnamed:') else col for col in df.columns]
                
                # Remove columns that are completely unnamed or have generic names
                df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
                
                # If the dataframe has data after cleaning, return it
                if not df.empty and len(df) > 0:
                    return df
            
            return None
            
        except Exception as e:
            print(f"Error reading Excel worksheet {worksheet_name} from {file_path}: {e}")
            return None
        finally:
            # Ensure Excel file is properly closed
            if excel_file is not None:
                try:
                    excel_file.close()
                except:
                    pass
    
    @staticmethod
    def _create_formatted_worksheet(wb: Workbook, worksheet_name: str, data: pd.DataFrame, config: Dict[str, str]):
        """Create professionally formatted worksheet with data"""
        ws = wb.create_sheet(worksheet_name)
        
        # Define professional styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue header
        alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray alternating rows
        title_font = Font(color="000000", bold=True, size=14)
        subtitle_font = Font(color="000000", bold=True, size=12)
        header_font = Font(color="FFFFFF", bold=True, size=11)  # White bold text for headers
        normal_font = Font(color="000000", size=10)  # Black text for data
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        # Add title
        title = f"Goods and Services Tax - Form GSTR-1"
        ws.cell(row=1, column=1, value=title)
        ws.merge_cells('A1:P1')
        title_cell = ws['A1']
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        
        # Add subtitle
        subtitle = f"{config.get('table_ref', '')} - {config['reference']}"
        ws.cell(row=3, column=1, value=subtitle)
        ws.merge_cells('A3:P3')
        subtitle_cell = ws['A3']
        subtitle_cell.font = subtitle_font
        subtitle_cell.alignment = left_alignment
        
        # Add data starting from row 5
        if not data.empty:
            # Add headers with professional styling
            header_row = 5
            for col, column_name in enumerate(data.columns, 1):
                cell = ws.cell(row=header_row, column=col, value=column_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Add data rows with alternating colors
            for row_idx, (_, row_data) in enumerate(data.iterrows()):
                excel_row = header_row + 1 + row_idx
                use_alt_color = (row_idx % 2 == 1)  # Alternate every other row
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=excel_row, column=col_idx, value=value)
                    cell.font = normal_font
                    cell.border = border
                    
                    # Apply alignment based on data type
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = left_alignment
                    
                    # Apply alternating row colors
                    if use_alt_color:
                        cell.fill = alt_row_fill
            
            # Auto-adjust column widths (safer approach avoiding merged cells)
            if not data.empty:
                # Calculate column widths based on data content
                for col_idx, column_name in enumerate(data.columns, 1):
                    max_length = len(str(column_name))  # Start with header length
                    
                    # Check data in this column
                    for value in data.iloc[:, col_idx-1]:  # Use iloc to access by position
                        try:
                            if pd.notna(value):
                                value_length = len(str(value))
                                if value_length > max_length:
                                    max_length = value_length
                        except Exception:
                            pass
                    
                    # Set width with reasonable limits
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    
                    # Get column letter for the current column
                    column_letter = get_column_letter(col_idx)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
        else:
            # No data message with proper styling
            no_data_cell = ws.cell(row=5, column=1, value="No data available for this section")
            no_data_cell.font = Font(color="666666", italic=True, size=10)
            no_data_cell.alignment = center_alignment
    
    @staticmethod
    def _get_month_name(month: int, year: int) -> str:
        """Get month name in MMM-YY format"""
        month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        
        if 1 <= month <= 12:
            return f"{month_names[month-1]}-{str(year)[-2:]}"
        
        return f"Invalid-{year}"
    
    @staticmethod
    def _extract_period_from_filename(filename: str) -> str:
        """Extract period from filename and return in MMM-YY format"""
        import re
        
        # Try to find patterns like 042021, 052021, etc. (MMYYYY format)
        mmyyyy_pattern = r'(\d{2})(\d{4})'
        match = re.search(mmyyyy_pattern, filename)
        
        if match:
            month = int(match.group(1))
            year = int(match.group(2))
            return GSTAnalysisService._get_month_name(month, year)
        
        # Try to find patterns like 04-2021, 05-2021, etc.
        mmyyyy_dash_pattern = r'(\d{2})-(\d{4})'
        match = re.search(mmyyyy_dash_pattern, filename)
        
        if match:
            month = int(match.group(1))
            year = int(match.group(2))
            return GSTAnalysisService._get_month_name(month, year)
        
        # Fallback: return filename without extension
        return filename.replace('.xlsx', '').replace('.xls', '').replace('.zip', '')
    
    @staticmethod
    def _add_month_validation(data: pd.DataFrame, month_column: str, date_column: str) -> tuple[pd.DataFrame, int]:
        """Add month validation column and return count of wrong month records"""
        import re
        from datetime import datetime
        
        wrong_month_count = 0
        
        # Get the month from the Month column (already added by _read_worksheet_data)
        if 'Month' not in data.columns:
            return data, 0
        
        file_month = data['Month'].iloc[0] if len(data) > 0 else ""
        
        # Create validation column
        validation_results = []
        
        for idx, row in data.iterrows():
            try:
                # Get the date from the specified date column
                date_value = row.get(date_column, "")
                
                if pd.isna(date_value) or str(date_value).strip() == "":
                    validation_results.append("NO DATE")
                    continue
                
                # Try to parse the date
                date_str = str(date_value)
                
                # Common date patterns
                date_patterns = [
                    r'(\d{2})[/-](\d{2})[/-](\d{4})',  # DD/MM/YYYY or DD-MM-YYYY
                    r'(\d{4})[/-](\d{2})[/-](\d{2})',  # YYYY/MM/DD or YYYY-MM-DD
                    r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})',  # D/M/YYYY or DD/M/YYYY
                ]
                
                parsed_date = None
                for pattern in date_patterns:
                    match = re.search(pattern, date_str)
                    if match:
                        try:
                            groups = match.groups()
                            if len(groups) == 3:
                                # Try different date formats
                                if len(groups[0]) == 4:  # YYYY format
                                    year, month, day = int(groups[0]), int(groups[1]), int(groups[2])
                                else:  # DD/MM/YYYY format
                                    day, month, year = int(groups[0]), int(groups[1]), int(groups[2])
                                
                                parsed_date = datetime(year, month, day)
                                break
                        except (ValueError, TypeError):
                            continue
                
                if parsed_date:
                    # Get month from parsed date
                    data_month = GSTAnalysisService._get_month_name(parsed_date.month, parsed_date.year)
                    
                    if data_month == file_month:
                        validation_results.append("CORRECT MONTH")
                    else:
                        validation_results.append("WRONG MONTH")
                        wrong_month_count += 1
                else:
                    validation_results.append("INVALID DATE")
                    
            except Exception as e:
                validation_results.append("ERROR")
        
        # Add validation column at the end of the dataframe
        data['Month Validation'] = validation_results
        
        return data, wrong_month_count
